#!/usr/bin/env python3
"""
Flexible JSON to Excel converter
Each JSON file becomes one row in the Excel file. Nested fields are serialized as JSON strings.
"""

import typer
from pathlib import Path
import json
import pandas as pd
from typing import Any, Dict
from rich.console import Console
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
console = Console()

def flatten_json(y: Dict[str, Any], parent_key: str = '', sep: str = '.') -> Dict[str, Any]:
    """Flatten a nested json file, rendering lists as readable multi-line strings."""
    items = {}
    for k, v in y.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.update(flatten_json(v, new_key, sep=sep))
        elif isinstance(v, list):
            # Render lists of dicts as readable multi-line strings
            if all(isinstance(i, dict) for i in v):
                rendered = []
                for d in v:
                    # Render all key-value pairs, one per line
                    lines = [f"{str(key).replace('_', ' ').title()}: {val}" for key, val in d.items()]
                    rendered.append('\n'.join(lines))
                items[new_key] = '\n\n'.join(rendered)
            elif all(isinstance(i, str) for i in v):
                items[new_key] = '\n\n'.join(v)
            else:
                # Fallback: JSON string
                items[new_key] = json.dumps(v, ensure_ascii=False)
        else:
            items[new_key] = v
    return items

def process_json_files(source_folder: Path) -> pd.DataFrame:
    """Process all JSON files in the folder, flattening each to a row."""
    json_files = list(source_folder.rglob("*.json"))
    if not json_files:
        raise FileNotFoundError("No JSON files found in the source folder.")
    rows = []
    all_keys = set()
    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            flat = flatten_json(data)
            flat['__file__'] = str(json_file)
            rows.append(flat)
            all_keys.update(flat.keys())
        except Exception as e:
            logger.error(f"Failed to process {json_file}: {e}")
    # Ensure all rows have all columns
    all_keys = sorted(all_keys)
    normalized_rows = []
    for row in rows:
        normalized_rows.append({k: row.get(k, '') for k in all_keys})
    df = pd.DataFrame(normalized_rows)
    return df

def json_to_excel(
    source_folder: Path = typer.Argument(..., help="Source folder containing JSON files"),
    output_file: Path = typer.Argument(..., help="Output Excel file (e.g. output.xlsx)")
):
    """Convert all JSON files in a folder to a single Excel file, one row per file."""
    console.print(f"[blue]Converting JSON files to Excel document")
    console.print(f"[cyan]Source folder: {source_folder}")
    console.print(f"[cyan]Output file: {output_file}")
    try:
        df = process_json_files(source_folder)

        # Reorder columns: resume/summary, tags, key people first if present
        preferred = []
        col_map = {}
        for col in df.columns:
            col_lc = col.lower()
            if col_lc in ["resume", "resumen", "summary"]:
                preferred.append(col)
                col_map["summary"] = col
            elif col_lc in ["tags", "etiquetas"]:
                preferred.append(col)
                col_map["tags"] = col
            elif col_lc in ["key people", "personas_clave", "personas_clave_y_etiquetas", "key_people", "key_people_and_tags"]:
                preferred.append(col)
                col_map["key_people"] = col
        # Remove duplicates
        preferred = list(dict.fromkeys(preferred))
        # Add the rest
        rest = [c for c in df.columns if c not in preferred]
        new_order = preferred + rest
        df = df[new_order]

        df.to_excel(output_file, index=False, engine='openpyxl')

        # Style the Excel file
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
        wb = load_workbook(output_file)
        ws = wb.active
        # Set header style
        for cell in ws[1]:
            cell.font = Font(name='Helvetica', size=10, bold=True)
            # Title case for header
            cell.value = str(cell.value).replace('_', ' ').title()
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        # Set all other cells font, word wrap, left/top alignment
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Helvetica', size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        wb.save(output_file)
        console.print(f"[green]Successfully wrote Excel file: {output_file}")
    except Exception as e:
        console.print(f"[red]Failed to convert JSON files: {e}")

if __name__ == "__main__":
    typer.run(json_to_excel) 