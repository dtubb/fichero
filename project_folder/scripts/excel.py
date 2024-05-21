import typer
import lxml.etree as etree
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from rich.progress import Progress

app = typer.Typer()

def process_xml(xml_file: Path, force: bool = False) -> dict:
    """
    Process an XML file to extract relevant data.

    :param xml_file: Path to the XML file
    :param force: Flag to force reprocessing of the Excel file
    :return: Dictionary containing the extracted data
    """
    
    try:
        # Load XML with recovery option
        parser = etree.XMLParser(recover=True)
        tree = etree.parse(xml_file, parser)
    except Exception as e:
        print(f"Error parsing {xml_file.name}: {e}")
        return None

    # Define the namespace
    ns = {'alto': 'http://www.loc.gov/standards/alto/ns-v4#'}

    # Find Page element
    page_elem = tree.find(".//{http://www.loc.gov/standards/alto/ns-v4#}Page")

    if page_elem is None:
        print(f"No Page element found in {xml_file.name}")
        return None

    # Get the String element under the Page element
    string_elem = page_elem.find("{http://www.loc.gov/standards/alto/ns-v4#}String")

    if string_elem is None:
        print(f"No String element found under Page in {xml_file.name}")
        return None

    # Get the full text from the String element
    full_text = string_elem.get("CONTENT", "").replace('\n', '\n')  # Add two lines after each line

    # Get the page entities
    entities = {}
    for entity_elem in page_elem.findall(".//{http://www.loc.gov/standards/alto/ns-v4#}Entity"):
        entity_type = entity_elem.get("type")
        entity_text = entity_elem.text.replace('\n', ' ').strip()  # Strip newlines and leading/trailing whitespace
        entities.setdefault(entity_type, []).append(entity_text)

    # Get the LLM catalogue entries
    catalogue_entries = {}
    llm_catalogue_elem = page_elem.find(".//{http://www.loc.gov/standards/alto/ns-v4#}LLMCatalogue")
    if llm_catalogue_elem is not None:
        for category_elem in llm_catalogue_elem:
            category_name = category_elem.tag.split('}')[-1]
            for result_elem in category_elem.findall(".//{http://www.loc.gov/standards/alto/ns-v4#}Result"):
                if result_elem.text is not None:
                    result_text = result_elem.text.replace('\n', ' ').strip()  # Strip newlines and leading/trailing whitespace
                    catalogue_entries.setdefault(category_name, []).append(result_text)

    # Create a dictionary with the extracted data
    data = {
        'File Name': xml_file.name,
        'Markdown File': f'=HYPERLINK("{xml_file.parent / f"{xml_file.stem}.md"}","MD")',
        'Image File': f'=HYPERLINK("{xml_file.parent / xml_file.with_suffix(".jpg")}","JPG")',
        'Summary': catalogue_entries.get('Summary', [''])[0],
        'People': ' ; '.join(list(set(entities.get('PER', []) + entities.get('People', [])))),
        'Organizations': ' ; '.join(list(set(entities.get('ORG', []) + entities.get('Organizations', [])))),
        'Places': ' ; '.join(list(set(entities.get('LOC', []) + entities.get('Places', [])))),
        'Keywords': ' ; '.join(list(set(entities.get('MISC', []) + entities.get('Keywords', [])))),
        'Full Text': full_text
    }

    return data

@app.command()
def main(
    folder_path: Path = typer.Argument(..., help="Path to the folder containing XML files"),
    force: bool = typer.Option(False, help="Force reprocessing of the Excel file")
):
    """
    Process XML files in a folder to extract relevant data and create an Excel spreadsheet.
    """

    # Find all XML files and sort them
    xml_files = []
    extensions = ["xml"]
    for ext in extensions:
        xml_files.extend(folder_path.rglob(f"**/*.{ext}"))

    xml_files = sorted(xml_files)

    total_files = len(xml_files)
    print(f"Processing {total_files} files...")

    # Get the parent folder name for the output file
    output_file = folder_path / f"{folder_path.name}-catalogue.xlsx"

    # Check if the output file already exists and skip if not forced
    if output_file.exists() and not force:
        print(f"Excel file '{output_file.name}' already exists. Use the --force option to overwrite.")
        return

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the header row
    header = ['File Name', 'Markdown File', 'Image File', 'Summary', 'People', 'Organizations', 'Places', 'Keywords', 'Full Text']
    sheet.append(header)

    # Set column widths and styles
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 5
    sheet.column_dimensions['C'].width = 5
    sheet.column_dimensions['D'].width = 60
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 30
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 90
    
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    # Bold the header row
    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)

    with Progress() as progress:
        task = progress.add_task("[green]Processing...", total=total_files)
        processed_files_count = 0

        for xml_file in xml_files:
            data = process_xml(xml_file, force)
            if data:
                # Write the data to the sheet
                row = [data[key] for key in header]
                sheet.append(row)

                # Set row height to auto
                row_num = sheet.max_row
                sheet.row_dimensions[row_num].auto_size = True
                processed_files_count += 1
            progress.update(task, advance=1)

    # Set font size for remaining rows
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = Font(size=14)
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

    # Save the workbook
    workbook.save(output_file)

    print(f"Processed {processed_files_count} files.")
    print(f"Catalogue saved to {output_file}")

if __name__ == "__main__":
    app()